import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
from flask import current_app
import logging
from models import db, Student, Teacher, Course, Score, CourseRegistration, Subject

logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_font = Font(color='FFFFFF', bold=True)
        self.center_align = Alignment(horizontal='center', vertical='center')

    def export_students_list(self, students, output_path=None):
        """Export students list to Excel"""
        if not output_path:
            output_path = f'students_list_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách Sinh viên"
        
        # Headers
        headers = ['STT', 'Mã SV', 'Họ và tên', 'Lớp', 'Khóa', 'Email', 'Số điện thoại', 'Ngày sinh', 'Giới tính', 'Trạng thái', 'GPA']
        ws.append(headers)
        
        # Apply styles to headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        
        # Data
        for i, student in enumerate(students, 1):
            ws.append([
                i,
                student.student_id,
                student.user.full_name,
                student.class_.class_name if student.class_ else 'N/A',
                student.course,
                student.user.email,
                student.user.phone or '',
                student.birth_date.strftime('%d/%m/%Y') if student.birth_date else '',
                student.gender or '',
                student.status,
                student.gpa
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        try:
            wb.save(output_path)
            return output_path
        except Exception as e:
            logger.error(f"Error exporting students list to Excel: {e}")
            raise

    def export_scores_sheet(self, course, scores, output_path=None):
        """Export course scores to Excel"""
        if not output_path:
            output_path = f'scores_{course.course_code}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bảng điểm"
        
        # Course information
        ws.append(['Mã học phần:', course.course_code])
        ws.append(['Tên học phần:', course.subject.subject_name])
        ws.append(['Giảng viên:', course.teacher.user.full_name])
        ws.append(['Học kỳ:', f"HK{course.semester} - {course.year}"])
        ws.append(['Số tín chỉ:', course.subject.credits])
        ws.append([])
        
        # Scores headers
        headers = ['STT', 'Mã SV', 'Họ và tên', 'Lớp']
        
        # Add grading components headers
        grading_components = []
        if course.grading_components:
            try:
                import json
                components = json.loads(course.grading_components)
                for component in components:
                    headers.append(f"{component['name']} ({component['weight']}%)")
                    grading_components.append(component)
            except:
                pass
        
        headers.extend(['Điểm QT', 'Điểm thi', 'Điểm tổng', 'Điểm chữ', 'Kết quả', 'Ghi chú'])
        ws.append(headers)
        
        # Apply styles to headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=7, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        
        # Scores data
        for i, score in enumerate(scores, 1):
            row_data = [
                i,
                score.student.student_id,
                score.student.user.full_name,
                score.student.class_.class_name if score.student.class_ else 'N/A'
            ]
            
            # Add component scores
            if score.components:
                try:
                    components_data = json.loads(score.components)
                    for component in grading_components:
                        comp_score = components_data.get(component['name'], '')
                        row_data.append(comp_score)
                except:
                    for _ in grading_components:
                        row_data.append('')
            else:
                for _ in grading_components:
                    row_data.append('')
            
            row_data.extend([
                score.process_score or '',
                score.exam_score or '',
                score.final_score or '',
                score.grade or '',
                'Đạt' if score.final_score and score.final_score >= 5.0 else 'Trượt',
                score.notes or ''
            ])
            
            ws.append(row_data)
        
        # Statistics
        ws.append([])
        passed_count = len([s for s in scores if s.final_score and s.final_score >= 5.0])
        total_count = len(scores)
        pass_rate = (passed_count / total_count * 100) if total_count > 0 else 0
        
        ws.append(['Thống kê:'])
        ws.append(['Tổng số sinh viên:', total_count])
        ws.append(['Số sinh viên đạt:', passed_count])
        ws.append(['Tỷ lệ đạt:', f"{pass_rate:.1f}%"])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        try:
            wb.save(output_path)
            return output_path
        except Exception as e:
            logger.error(f"Error exporting scores to Excel: {e}")
            raise

    def export_course_registrations(self, registrations, output_path=None):
        """Export course registrations to Excel"""
        if not output_path:
            output_path = f'course_registrations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Đăng ký học phần"
        
        # Headers
        headers = ['STT', 'Mã SV', 'Họ và tên', 'Lớp', 'Mã HP', 'Tên học phần', 'Số TC', 'Giảng viên', 'Ngày đăng ký', 'Trạng thái', 'Ghi chú']
        ws.append(headers)
        
        # Apply styles to headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        
        # Data
        for i, reg in enumerate(registrations, 1):
            ws.append([
                i,
                reg.student.student_id,
                reg.student.user.full_name,
                reg.student.class_.class_name if reg.student.class_ else 'N/A',
                reg.course.course_code,
                reg.course.subject.subject_name,
                reg.course.subject.credits,
                reg.course.teacher.user.full_name,
                reg.registration_date.strftime('%d/%m/%Y %H:%M'),
                reg.status,
                reg.notes or ''
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        try:
            wb.save(output_path)
            return output_path
        except Exception as e:
            logger.error(f"Error exporting course registrations to Excel: {e}")
            raise

    def export_teaching_schedule(self, teacher, courses, output_path=None):
        """Export teaching schedule to Excel"""
        if not output_path:
            output_path = f'teaching_schedule_{teacher.teacher_code}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Lịch giảng dạy"
        
        # Teacher information
        ws.append(['Mã giáo viên:', teacher.teacher_code])
        ws.append(['Họ và tên:', teacher.user.full_name])
        ws.append(['Bộ môn:', teacher.department])
        ws.append(['Thời kỳ:', f"HK{courses[0].semester if courses else ''} - {courses[0].year if courses else ''}"])
        ws.append([])
        
        # Headers
        headers = ['STT', 'Mã HP', 'Tên học phần', 'Số TC', 'Số SV', 'Lịch học', 'Phòng', 'Trạng thái']
        ws.append(headers)
        
        # Apply styles to headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=7, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        
        # Data
        for i, course in enumerate(courses, 1):
            ws.append([
                i,
                course.course_code,
                course.subject.subject_name,
                course.subject.credits,
                course.current_students,
                course.schedule or 'N/A',
                course.room or 'N/A',
                course.status
            ])
        
        # Summary
        total_credits = sum(course.subject.credits for course in courses)
        total_students = sum(course.current_students for course in courses)
        
        ws.append([])
        ws.append(['Tổng số học phần:', len(courses)])
        ws.append(['Tổng số tín chỉ:', total_credits])
        ws.append(['Tổng số sinh viên:', total_students])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        try:
            wb.save(output_path)
            return output_path
        except Exception as e:
            logger.error(f"Error exporting teaching schedule to Excel: {e}")
            raise

    def import_students_from_excel(self, file_path):
        """Import students from Excel file"""
        try:
            df = pd.read_excel(file_path)
            imported_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Validate required fields
                    required_fields = ['student_id', 'full_name', 'email', 'course']
                    for field in required_fields:
                        if pd.isna(row.get(field)):
                            errors.append(f"Dòng {index + 2}: Thiếu trường {field}")
                            continue
                    
                    # Check if student already exists
                    existing_student = Student.query.filter_by(student_id=row['student_id']).first()
                    if existing_student:
                        errors.append(f"Dòng {index + 2}: Mã SV {row['student_id']} đã tồn tại")
                        continue
                    
                    # Check if user already exists
                    existing_user = User.query.filter_by(email=row['email']).first()
                    if existing_user:
                        errors.append(f"Dòng {index + 2}: Email {row['email']} đã tồn tại")
                        continue
                    
                    # Create user
                    user = User(
                        username=row['student_id'],
                        email=row['email'],
                        full_name=row['full_name'],
                        role=UserRole.STUDENT,
                        phone=str(row.get('phone', '')),
                        address=row.get('address', '')
                    )
                    user.set_password('123456')  # Default password
                    
                    # Create student
                    student = Student(
                        user=user,
                        student_id=row['student_id'],
                        course=row['course'],
                        birth_date=pd.to_datetime(row['birth_date']) if pd.notna(row.get('birth_date')) else None,
                        gender=row.get('gender', '')
                    )
                    
                    db.session.add(user)
                    db.session.add(student)
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Dòng {index + 2}: {str(e)}")
                    continue
            
            if imported_count > 0:
                db.session.commit()
            
            return imported_count, errors
            
        except Exception as e:
            logger.error(f"Error importing students from Excel: {e}")
            raise

    def import_scores_from_excel(self, file_path, course_id):
        """Import scores from Excel file"""
        try:
            df = pd.read_excel(file_path)
            updated_count = 0
            errors = []
            
            course = Course.query.get(course_id)
            if not course:
                raise ValueError("Course not found")
            
            for index, row in df.iterrows():
                try:
                    student_id = row['student_id']
                    student = Student.query.filter_by(student_id=student_id).first()
                    
                    if not student:
                        errors.append(f"Dòng {index + 2}: Không tìm thấy SV {student_id}")
                        continue
                    
                    # Find or create score record
                    score = Score.query.filter_by(student_id=student.id, course_id=course_id).first()
                    if not score:
                        score = Score(student_id=student.id, course_id=course_id)
                    
                    # Update scores
                    if pd.notna(row.get('process_score')):
                        score.process_score = float(row['process_score'])
                    if pd.notna(row.get('exam_score')):
                        score.exam_score = float(row['exam_score'])
                    if pd.notna(row.get('final_score')):
                        score.final_score = float(row['final_score'])
                    
                    score.notes = row.get('notes', '')
                    
                    if not score.id:
                        db.session.add(score)
                    
                    updated_count += 1
                    
                except Exception as e:
                    errors.append(f"Dòng {index + 2}: {str(e)}")
                    continue
            
            if updated_count > 0:
                db.session.commit()
            
            return updated_count, errors
            
        except Exception as e:
            logger.error(f"Error importing scores from Excel: {e}")
            raise

# Utility functions for easy access
def export_to_excel(data_type, data, output_path=None):
    """Utility function to export different types of data to Excel"""
    generator = ExcelGenerator()
    
    if data_type == 'students':
        return generator.export_students_list(data, output_path)
    elif data_type == 'scores':
        return generator.export_scores_sheet(data['course'], data['scores'], output_path)
    elif data_type == 'registrations':
        return generator.export_course_registrations(data, output_path)
    elif data_type == 'teaching_schedule':
        return generator.export_teaching_schedule(data['teacher'], data['courses'], output_path)
    else:
        raise ValueError(f"Unsupported data type: {data_type}")

def import_from_excel(file_type, file_path, **kwargs):
    """Utility function to import data from Excel"""
    generator = ExcelGenerator()
    
    if file_type == 'students':
        return generator.import_students_from_excel(file_path)
    elif file_type == 'scores':
        return generator.import_scores_from_excel(file_path, kwargs.get('course_id'))
    else:
        raise ValueError(f"Unsupported file type: {file_type}")